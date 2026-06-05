from pathlib import Path

import openpyxl


path = Path("WEBFS - Nieuwe Functionaliteit v2.xlsx")
wb = openpyxl.load_workbook(path, data_only=True)
print(wb.sheetnames)
for ws in wb.worksheets:
    print(f"SHEET\t{ws.title}\trows={ws.max_row}\tcols={ws.max_column}")
    for row in ws.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if any(values):
            print("\t".join(values))
    print("---")
